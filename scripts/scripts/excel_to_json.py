"""
엑셀 → JSON 변환 스크립트

1) 월별손익.xlsx → pl_monthly.json
2) S&C raw.xlsx  → sales_cost_summary.json, sales_cost_raw.json
3) RAW(E)_XXXX.xlsx → expense_detail.json
4) CH_VC_in.xlsx → kpi_ch.json
5) 건기식_직영몰,다이소매출.xlsx → kpi_health.json
6) 나보타_계약,발매국가수.xlsx → kpi_nabota.json
"""

import json
import math
import os
import sys
from collections import defaultdict

import openpyxl

# ─── 경로 설정 ───
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_PL = os.path.join(BASE_DIR, "input", "월별손익.xlsx")
INPUT_SC = os.path.join(BASE_DIR, "input", "S&C raw.xlsx")
INPUT_RAWE_PATTERN = os.path.join(BASE_DIR, "input", "RAW(E)_{year}.xlsx")
RAWE_YEARS = [2022, 2023, 2024, 2025]
OUTPUT_DIR = os.path.join(BASE_DIR, "public", "data")
OUTPUT_PL = os.path.join(OUTPUT_DIR, "pl_monthly.json")
OUTPUT_SC_SUMMARY = os.path.join(OUTPUT_DIR, "sales_cost_summary.json")
OUTPUT_SC_RAW = os.path.join(OUTPUT_DIR, "sales_cost_raw.json")
OUTPUT_EXPENSE = os.path.join(OUTPUT_DIR, "expense_detail.json")

# ─── KPI 추가 데이터 ───
INPUT_KPI_CH = os.path.join(BASE_DIR, "input", "CH_VC_in.xlsx")
INPUT_KPI_HEALTH = os.path.join(BASE_DIR, "input", "건기식_직영몰,다이소매출.xlsx")
INPUT_KPI_NABOTA = os.path.join(BASE_DIR, "input", "나보타_계약,발매국가수.xlsx")
OUTPUT_KPI_CH = os.path.join(OUTPUT_DIR, "kpi_ch.json")
OUTPUT_KPI_HEALTH = os.path.join(OUTPUT_DIR, "kpi_health.json")
OUTPUT_KPI_NABOTA = os.path.join(OUTPUT_DIR, "kpi_nabota.json")

# ─── 카테고리 더미값 (null 처리 대상) ───
DUMMY_VALUES = {"'0", "(비어 있음)", "0", ""}

# ─── 비용분류No → JSON 키 매핑 (44개 항목) ───
ITEM_MAP = {
    1: "매출",
    2: "매출_국내",
    3: "매출_수출",
    4: "매출원가",
    5: "폐기비용",
    6: "매출총이익",
    7: "영업판관비",
    8: "영업판관비_영업비용",
    9: "영업판관비_마케팅비용",
    10: "영업판관비_영업직접비",
    11: "영업판관비_영업인건비",
    12: "영업판관비_마케팅인건비",
    13: "영업판관비_광고비",
    14: "판매대행수수료",
    15: "판매대행수수료_국내",
    16: "판매대행수수료_해외",
    17: "매출변동비",
    18: "매출변동비_운반비",
    19: "매출변동비_쇼핑몰수수료",
    20: "매출변동비_OTC로열티",
    21: "매출변동비_ETC로열티",
    22: "매출변동비_EGF로열티",
    23: "매출변동비_카드수수료",
    24: "영업관리비",
    25: "영업관리비_인건비",
    26: "영업관리비_지사운영비",
    27: "영업관리비_감가상각비",
    28: "영업관리비_기타경비",
    29: "일반관리비",
    30: "일반관리비_인건비",
    31: "일반관리비_대웅용역료",
    32: "일반관리비_감가상각비",
    33: "일반관리비_IT비용",
    34: "일반관리비_세금과공과",
    35: "일반관리비_지급수수료",
    36: "일반관리비_기타경비",
    37: "비효율비경상비용",
    38: "비효율비경상비용_소송비용",
    39: "비효율비경상비용_대손상각비",
    40: "R&D차감전이익",
    41: "R&D비용",
    42: "R&D비용_R연구",
    43: "R&D비용_D개발",
    44: "영업이익",
}

# ─── 사업부 컬럼 매핑 (openpyxl 1-indexed) ───
# (사업부명, 금액컬럼, %컬럼)
DIVISIONS = [
    ("ETC", 13, 14),
    ("CH", 15, 16),
    ("건기식", 17, 18),
    ("나보타", 19, 20),
    ("글로벌", 21, 22),
    ("수탁", 23, 24),
    ("기타", 25, 26),
    ("전사", 27, 28),
]


def clean_value(v):
    """셀 값 정리: NaN/None/빈값 → None, 숫자는 round(2)"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "'0", "0", "-"):
            return None
        try:
            v = float(v)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return round(v, 2)
    return None


def clean_rate(v):
    """비율 값 정리: NaN/None → None, 숫자는 round(4)"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "'0", "0", "-"):
            return None
        try:
            v = float(v)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return round(v, 4)
    return None


def parse_sheet(ws, sheet_type):
    """시트에서 월별 손익 데이터를 파싱하여 리스트로 반환"""
    # (type, year, month) → {division: {item_key: value}}
    records = defaultdict(lambda: defaultdict(dict))
    quarters = {}

    for row_idx in range(2, ws.max_row + 1):
        year_val = ws.cell(row=row_idx, column=3).value
        month_val = ws.cell(row=row_idx, column=4).value
        quarter_val = ws.cell(row=row_idx, column=5).value
        item_no_val = ws.cell(row=row_idx, column=6).value

        # 유효한 데이터 행인지 확인
        if year_val is None or month_val is None or month_val == "":
            continue
        try:
            year = int(year_val)
            month = int(month_val)
        except (ValueError, TypeError):
            continue

        if item_no_val is None or item_no_val == "":
            continue
        try:
            item_no = int(item_no_val)
        except (ValueError, TypeError):
            continue

        if item_no not in ITEM_MAP:
            continue

        item_key = ITEM_MAP[item_no]
        rec_key = (sheet_type, year, month)

        # 분기 저장
        if quarter_val is not None and quarter_val != "":
            try:
                quarters[rec_key] = int(quarter_val)
            except (ValueError, TypeError):
                pass

        # 각 사업부별 금액/비율 추출
        for div_name, amt_col, pct_col in DIVISIONS:
            amt = clean_value(ws.cell(row=row_idx, column=amt_col).value)
            pct = clean_rate(ws.cell(row=row_idx, column=pct_col).value)

            if amt is not None:
                records[rec_key][div_name][item_key] = amt
            else:
                records[rec_key][div_name][item_key] = None

            # 비율은 매출원가(No.4)만 저장 (매출원가율)
            if item_no == 4 and pct is not None:
                records[rec_key][div_name]["매출원가율"] = pct

    # 레코드를 리스트로 변환
    result = []
    for (typ, year, month), divisions in sorted(records.items()):
        quarter = quarters.get((typ, year, month), (month - 1) // 3 + 1)
        entry = {
            "type": typ,
            "year": year,
            "month": month,
            "quarter": quarter,
            "items": {},
        }
        for div_name, _, _ in DIVISIONS:
            if div_name in divisions:
                entry["items"][div_name] = divisions[div_name]
            else:
                entry["items"][div_name] = {}
        result.append(entry)

    return result


def convert_pl_monthly():
    """월별손익.xlsx → pl_monthly.json 변환"""
    print(f"[1/3] 엑셀 파일 로드: {INPUT_PL}")
    wb = openpyxl.load_workbook(INPUT_PL, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  시트 목록: {sheet_names}")

    # 실적 시트 파싱
    ws_actual = wb[sheet_names[0]]  # "실적"
    print(f"[2/3] 실적 시트 파싱 ({ws_actual.max_row} rows)...")
    actual_data = parse_sheet(ws_actual, "실적")
    print(f"  → {len(actual_data)}개 월별 레코드")

    # 26년목표 시트 파싱
    ws_target = wb[sheet_names[1]]  # "26년목표"
    print(f"  26년목표 시트 파싱 ({ws_target.max_row} rows)...")
    target_data = parse_sheet(ws_target, "목표")
    print(f"  → {len(target_data)}개 월별 레코드")

    # 결과 조합
    all_data = actual_data + target_data

    # 연도 목록 추출
    years_actual = sorted(set(d["year"] for d in actual_data))
    years_target = sorted(set(d["year"] for d in target_data))

    output = {
        "metadata": {
            "unit": "억원",
            "lastUpdated": "2026-03-18",
            "periods_actual": [str(y) for y in years_actual],
            "periods_target": [str(y) for y in years_target],
        },
        "data": all_data,
    }

    # 출력
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_PL, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"[3/3] 저장 완료: {OUTPUT_PL}")
    print(f"  총 {len(all_data)}개 레코드 (실적 {len(actual_data)} + 목표 {len(target_data)})")

    return output


# ═══════════════════════════════════════════════════════════════
# S&C raw.xlsx 변환
# ═══════════════════════════════════════════════════════════════

# 헤더 반복 행 스킵용 (기획집계 컬럼에 이 값이 들어있으면 헤더 행)
SC_HEADER_VALUES = {"기획집계"}


def clean_category(v):
    """카테고리 컬럼 정리: 더미값 → None, 유효한 문자열은 그대로"""
    if v is None:
        return None
    s = str(v).strip()
    if s in DUMMY_VALUES:
        return None
    return s


def clean_amount_raw(v):
    """금액 값 정리 (원 단위 — 반올림 없이 정수/실수 유지)"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "-"):
            return None
        try:
            v = float(v)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        # 원 단위이므로 정수로 반환 (소수점 이하 반올림)
        return round(v)
    return None


def parse_sc_sheet(ws, has_product_code=False):
    """S&C raw 시트 파싱

    요약 시트: 11개 컬럼 (품목코드 없음)
    RAW 시트:  12개 컬럼 (col 10 = 품목코드)

    Returns: list of dict
    """
    records = []

    # 컬럼 인덱스 설정
    col_year = 1
    col_quarter = 2
    col_month = 3
    col_division = 4      # 기획집계
    col_product_type = 5   # 제상
    col_category = 6       # 중분류
    col_profit_tier = 7    # 수익군
    col_product_group = 8  # 품목구분1
    col_product_name = 9   # 품목구분2

    if has_product_code:
        col_product_code = 10
        col_sales = 11
        col_cost = 12
    else:
        col_product_code = None
        col_sales = 10
        col_cost = 11

    for row_idx in range(2, ws.max_row + 1):
        # 연도 유효성 체크
        year_val = ws.cell(row=row_idx, column=col_year).value
        if year_val is None:
            continue
        try:
            year = int(year_val)
        except (ValueError, TypeError):
            continue

        # 기획집계 — 헤더 반복 행 스킵
        division = ws.cell(row=row_idx, column=col_division).value
        if division is None:
            continue
        div_str = str(division).strip()
        if div_str in SC_HEADER_VALUES:
            continue

        # 월/분기
        month_val = ws.cell(row=row_idx, column=col_month).value
        quarter_val = ws.cell(row=row_idx, column=col_quarter).value
        try:
            month = int(month_val)
            quarter = int(quarter_val)
        except (ValueError, TypeError):
            continue

        # 카테고리 컬럼 정리
        division_clean = clean_category(division)
        product_type = clean_category(ws.cell(row=row_idx, column=col_product_type).value)
        category = clean_category(ws.cell(row=row_idx, column=col_category).value)
        profit_tier = clean_category(ws.cell(row=row_idx, column=col_profit_tier).value)
        product_group = clean_category(ws.cell(row=row_idx, column=col_product_group).value)
        product_name = clean_category(ws.cell(row=row_idx, column=col_product_name).value)

        # (비어 있음) 기획집계 → 스킵
        if division_clean is None:
            continue

        # 금액
        sales = clean_amount_raw(ws.cell(row=row_idx, column=col_sales).value)
        cost = clean_amount_raw(ws.cell(row=row_idx, column=col_cost).value)

        rec = {
            "year": year,
            "quarter": quarter,
            "month": month,
            "division": division_clean,
            "productType": product_type,
            "category": category,
            "profitTier": profit_tier,
            "productGroup": product_group,
            "productName": product_name,
            "sales": sales,
            "cost": cost,
        }

        if has_product_code:
            code_val = ws.cell(row=row_idx, column=col_product_code).value
            rec["productCode"] = clean_category(code_val)

        records.append(rec)

    return records


def convert_sales_cost():
    """S&C raw.xlsx → sales_cost_summary.json, sales_cost_raw.json 변환"""
    print(f"\n{'='*60}")
    print(f"S&C raw.xlsx 변환 시작")
    print(f"{'='*60}")
    print(f"[1/4] 엑셀 파일 로드: {INPUT_SC}")
    wb = openpyxl.load_workbook(INPUT_SC, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  시트 목록: {sheet_names}")

    # 요약 시트 파싱
    ws_summary = wb[sheet_names[0]]
    print(f"[2/4] 요약 시트 파싱 ({ws_summary.max_row} rows)...")
    summary_data = parse_sc_sheet(ws_summary, has_product_code=False)
    print(f"  -> {len(summary_data)}개 레코드")

    # RAW 시트 파싱
    ws_raw = wb[sheet_names[1]]
    print(f"[3/4] RAW 시트 파싱 ({ws_raw.max_row} rows)...")
    raw_data = parse_sc_sheet(ws_raw, has_product_code=True)
    print(f"  -> {len(raw_data)}개 레코드")

    # 연도 범위
    years_summary = sorted(set(d["year"] for d in summary_data))
    years_raw = sorted(set(d["year"] for d in raw_data))

    # sales_cost_summary.json 저장
    summary_output = {
        "metadata": {
            "unit": "원",
            "source": "S&C raw.xlsx 요약 시트",
            "lastUpdated": "2026-03-18",
            "years": years_summary,
        },
        "data": summary_data,
    }
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_SC_SUMMARY, "w", encoding="utf-8") as f:
        json.dump(summary_output, f, ensure_ascii=False, indent=2)
    print(f"  저장: {OUTPUT_SC_SUMMARY}")

    # sales_cost_raw.json 저장
    raw_output = {
        "metadata": {
            "unit": "원",
            "source": "S&C raw.xlsx RAW 시트",
            "lastUpdated": "2026-03-18",
            "years": years_raw,
        },
        "data": raw_data,
    }
    with open(OUTPUT_SC_RAW, "w", encoding="utf-8") as f:
        json.dump(raw_output, f, ensure_ascii=False, indent=2)
    print(f"  저장: {OUTPUT_SC_RAW}")

    print(f"[4/4] 완료 (요약 {len(summary_data)} + RAW {len(raw_data)} 레코드)")

    return summary_output, raw_output


# ═══════════════════════════════════════════════════════════════
# RAW(E) 변환
# ═══════════════════════════════════════════════════════════════

# RAW(E) 컬럼 매핑 (1-indexed, 모든 연도 공통)
RAWE_COL_CAT1 = 2       # 구분1
RAWE_COL_CAT2 = 3       # 구분2
RAWE_COL_CAT3 = 4       # 구분3
RAWE_COL_CAT4 = 5       # 구분4
RAWE_COL_BIZUNIT = 6    # 사업구분
RAWE_COL_MONTH_START = 11  # 월별 금액 시작 (col 11 = 1월, col 22 = 12월)


def parse_legal_cost_sheet(wb, year):
    """법무비용 시트 파싱 → (A열 구분, 월) 별 금액 합산

    A열='지급수수료' → 일반관리비-지급수수료로 재분류 필요
    A열='나보타'     → 비효율/비경상-소송비용에 유지 (변경 없음)

    Returns: dict  {(category_a, month): amount}
    """
    target = [s for s in wb.sheetnames if "법무" in s]
    if not target:
        return {}

    ws = wb[target[0]]
    result = defaultdict(int)

    for row_idx in range(2, ws.max_row + 1):
        cat_a = ws.cell(row=row_idx, column=1).value   # A열: 나보타/지급수수료
        period = ws.cell(row=row_idx, column=6).value   # F열: 기간(월)
        amount = ws.cell(row=row_idx, column=11).value  # K열: 금액

        if cat_a is None or period is None or amount is None:
            continue
        try:
            month = int(period)
            amt = int(round(float(str(amount))))
        except (ValueError, TypeError):
            continue

        result[(str(cat_a).strip(), month)] += amt

    return result


def apply_legal_cost_reclass(agg, legal_costs, year):
    """법무비용 재분류 적용

    지급수수료 금액을 (비효율/비경상비용, 소송비용) → (일반관리비, 지급수수료)로 이동
    """
    reclass_total = 0
    for (cat_a, month), amount in legal_costs.items():
        if cat_a != "지급수수료":
            continue
        if amount == 0:
            continue

        # 소송비용/법무비용/공통에서 차감
        src_key = ("비효율/비경상비용", "비효율/비경상비용", "소송비용", "법무비용", "공통", month)
        # 일반관리비/지급수수료/공통에 가산
        dst_key = ("공통비", "일반관리비", "지급수수료", "지급수수료", "공통", month)

        agg[src_key] -= amount
        agg[dst_key] += amount
        reclass_total += amount

    return reclass_total


def convert_expense_detail():
    """RAW(E)_2022~2025.xlsx → expense_detail.json 변환"""
    print(f"\n{'='*60}")
    print(f"RAW(E) 변환 시작")
    print(f"{'='*60}")

    all_records = []
    for year in RAWE_YEARS:
        fpath = INPUT_RAWE_PATTERN.format(year=year)
        print(f"  [{year}] 로드: {os.path.basename(fpath)}")
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb["RAW(E)"]
        print(f"    RAW(E): {ws.max_row} rows")

        # 1) RAW(E) 집계
        agg = defaultdict(int)
        for row_idx in range(2, ws.max_row + 1):
            cat1_raw = ws.cell(row=row_idx, column=RAWE_COL_CAT1).value
            if cat1_raw is None:
                continue
            cat1_str = str(cat1_raw).strip()
            if cat1_str == "제외1":
                continue

            cat1 = clean_category(cat1_raw)
            cat2 = clean_category(ws.cell(row=row_idx, column=RAWE_COL_CAT2).value)
            cat3 = clean_category(ws.cell(row=row_idx, column=RAWE_COL_CAT3).value)
            cat4 = clean_category(ws.cell(row=row_idx, column=RAWE_COL_CAT4).value)
            biz_raw = ws.cell(row=row_idx, column=RAWE_COL_BIZUNIT).value
            biz = clean_category(biz_raw)
            if biz is not None and biz == "제외1":
                continue

            for month in range(1, 13):
                col = RAWE_COL_MONTH_START + month - 1
                val = ws.cell(row=row_idx, column=col).value
                if val is None:
                    continue
                try:
                    amount = int(round(float(val)))
                except (ValueError, TypeError):
                    continue
                if amount == 0:
                    continue
                agg[(cat1, cat2, cat3, cat4, biz, month)] += amount

        # 2) 법무비용 재분류
        legal_costs = parse_legal_cost_sheet(wb, year)
        reclass_amt = apply_legal_cost_reclass(agg, legal_costs, year)
        if reclass_amt:
            print(f"    법무비용 재분류: {reclass_amt/1e8:.2f} 억원 (지급수수료 -> 일반관리비)")

        # 3) 집계 → 레코드 변환 (amount=0 제거)
        def sort_key(item):
            (c1, c2, c3, c4, b, m) = item[0]
            return (c1 or "", c2 or "", c3 or "", c4 or "", b or "", m)

        records = []
        for (cat1, cat2, cat3, cat4, biz, month), amount in sorted(agg.items(), key=sort_key):
            if amount == 0:
                continue
            records.append({
                "year": year,
                "month": month,
                "category1": cat1,
                "category2": cat2,
                "category3": cat3,
                "category4": cat4,
                "bizUnit": biz,
                "amount": amount,
            })

        print(f"    -> {len(records)}개 집계 레코드")
        all_records.extend(records)
        wb.close()

    output = {
        "metadata": {
            "unit": "원",
            "source": "RAW(E)_{연도}.xlsx RAW(E) 시트",
            "years": RAWE_YEARS,
            "lastUpdated": "2026-03-18",
        },
        "data": all_records,
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_EXPENSE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  저장: {OUTPUT_EXPENSE}")
    print(f"  총 {len(all_records)}개 레코드")
    return output


# ═══════════════════════════════════════════════════════════════
# 검증
# ═══════════════════════════════════════════════════════════════

def verify_pl(output):
    """2025년 1월 ETC 매출 713.28억원 검증"""
    print("\n=== pl_monthly 검증 ===")
    for entry in output["data"]:
        if entry["type"] == "실적" and entry["year"] == 2025 and entry["month"] == 1:
            etc_sales = entry["items"]["ETC"]["매출"]
            print(f"2025/01 ETC 매출: {etc_sales} 억원")
            if etc_sales == 713.28:
                print("[PASS] 713.28 억원 일치")
            else:
                print(f"[FAIL] 기대 713.28, 실제 {etc_sales}")
            return


def verify_sc(summary_output, pl_output):
    """S&C 요약 매출 합산 vs 월별손익 매출 교차 검증"""
    print("\n=== S&C 교차 검증 (2025/01 ETC) ===")

    # S&C 요약에서 2025/01 ETC 매출 합산
    sc_total = 0
    sc_count = 0
    for rec in summary_output["data"]:
        if rec["year"] == 2025 and rec["month"] == 1 and rec["division"] == "ETC":
            if rec["sales"] is not None:
                sc_total += rec["sales"]
                sc_count += 1
    sc_eok = round(sc_total / 1_0000_0000, 2)

    # 월별손익에서 2025/01 ETC 매출
    pl_eok = None
    for entry in pl_output["data"]:
        if entry["type"] == "실적" and entry["year"] == 2025 and entry["month"] == 1:
            pl_eok = entry["items"]["ETC"]["매출"]
            break

    print(f"  S&C 요약 합산: {sc_total:,} 원 = {sc_eok} 억원 ({sc_count}건)")
    print(f"  월별손익:       {pl_eok} 억원")

    if sc_eok == pl_eok:
        print(f"  [PASS] 일치 ({sc_eok} == {pl_eok})")
    else:
        diff = abs(sc_eok - pl_eok) if pl_eok else "N/A"
        print(f"  [WARN] 차이 {diff} 억원 (반올림 오차 가능)")

    # 전사 검증
    print("\n=== S&C 교차 검증 (2025/01 전사) ===")
    all_divisions = ["ETC", "CH", "건기식", "나보타", "글로벌", "수탁", "기타",
                     "매출할인", "원가추가"]
    div_totals = {}
    for div in all_divisions:
        total_s = 0
        total_c = 0
        cnt = 0
        for rec in summary_output["data"]:
            if rec["year"] == 2025 and rec["month"] == 1 and rec["division"] == div:
                if rec["sales"] is not None:
                    total_s += rec["sales"]
                if rec["cost"] is not None:
                    total_c += rec["cost"]
                cnt += 1
        if cnt > 0:
            div_totals[div] = (total_s, total_c, cnt)
            print(f"  {div}: 매출 {round(total_s/1e8, 2)} 억원, "
                  f"원가 {round(total_c/1e8, 2)} 억원 ({cnt}건)")

    grand_sales = sum(v[0] for v in div_totals.values())
    grand_cost = sum(v[1] for v in div_totals.values())
    print(f"  ---")
    print(f"  합계: 매출 {round(grand_sales/1e8, 2)} 억원, "
          f"원가 {round(grand_cost/1e8, 2)} 억원")

    # 월별손익 전사 매출과 비교
    for entry in pl_output["data"]:
        if entry["type"] == "실적" and entry["year"] == 2025 and entry["month"] == 1:
            pl_total_sales = entry["items"]["전사"]["매출"]
            pl_total_cost = entry["items"]["전사"]["매출원가"]
            print(f"  월별손익 전사: 매출 {pl_total_sales} 억원, "
                  f"매출원가 {pl_total_cost} 억원")
            break


def verify_expense(expense_output, pl_output):
    """RAW(E) 비용 합산 vs 월별손익 교차 검증"""
    print("\n=== RAW(E) 교차 검증 (2025/01 영업판관비) ===")

    # expense_detail에서 2025/01 구분2=영업판관비 합산
    total = 0
    count = 0
    by_cat3 = defaultdict(int)
    for rec in expense_output["data"]:
        if rec["year"] == 2025 and rec["month"] == 1 and rec["category2"] == "영업판관비":
            total += rec["amount"]
            count += 1
            by_cat3[rec["category3"]] += rec["amount"]

    eok = round(total / 1e8, 2)
    print(f"  RAW(E) 합산: {total:,} 원 = {eok} 억원 ({count}개 집계)")

    # 월별손익 전사 영업판관비
    pl_val = None
    for entry in pl_output["data"]:
        if entry["type"] == "실적" and entry["year"] == 2025 and entry["month"] == 1:
            pl_val = entry["items"]["전사"]["영업판관비"]
            break
    print(f"  월별손익 전사 영업판관비: {pl_val} 억원")

    diff = round(abs(eok - pl_val), 2) if pl_val else "N/A"
    if diff == 0:
        print(f"  [PASS] 정확히 일치")
    elif isinstance(diff, float) and diff < 1.0:
        print(f"  [PASS] 차이 {diff} 억원 (배부 조정 범위 내)")
    else:
        print(f"  [WARN] 차이 {diff} 억원")

    # 구분3별 내역
    print(f"\n  구분3별 내역:")
    for cat3, amt in sorted(by_cat3.items(), key=lambda x: -x[1]):
        print(f"    {cat3}: {round(amt/1e8, 2)} 억원")

    # 전체 비용 구분2별 합산
    print(f"\n=== RAW(E) 2025/01 구분2별 합산 ===")
    by_cat2 = defaultdict(int)
    for rec in expense_output["data"]:
        if rec["year"] == 2025 and rec["month"] == 1 and rec["category2"]:
            by_cat2[rec["category2"]] += rec["amount"]
    for cat2, amt in sorted(by_cat2.items(), key=lambda x: -abs(x[1])):
        print(f"  {cat2}: {round(amt/1e8, 2)} 억원")


def convert_kpi_simple(input_path, columns_config, output_path, description):
    """공통 KPI Excel → JSON 변환 함수.
    columns_config: list of (json_key, excel_col_index) pairs
    Excel 구조: 연도(A), 분기(B), 월(C), 데이터열(D~)
    """
    print(f"\n{'='*60}")
    print(f"  {description}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    output = {}
    for json_key, col_idx in columns_config:
        records = []
        for row in rows:
            year = row[0]
            quarter = row[1]
            month = row[2]
            value = row[col_idx]
            if year is None or month is None:
                continue
            records.append({
                "year": int(year),
                "quarter": int(quarter),
                "month": int(month),
                "value": int(value) if value is not None else None,
            })
        output[json_key] = records
        print(f"  {json_key}: {len(records)} records")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  → {output_path}")

    return output


def convert_kpi_ch():
    return convert_kpi_simple(
        INPUT_KPI_CH,
        [("vcInTypeCount", 3)],  # D열
        OUTPUT_KPI_CH,
        "CH VC인타입처수 변환",
    )


def convert_kpi_health():
    return convert_kpi_simple(
        INPUT_KPI_HEALTH,
        [("directMallSales", 3), ("daisoSales", 4)],  # D열, E열
        OUTPUT_KPI_HEALTH,
        "건기식 직영몰/다이소 매출 변환",
    )


def convert_kpi_nabota():
    return convert_kpi_simple(
        INPUT_KPI_NABOTA,
        [("contractCountries", 3), ("launchCountries", 4)],  # D열, E열
        OUTPUT_KPI_NABOTA,
        "나보타 계약/발매국가수 변환",
    )


if __name__ == "__main__":
    pl_output = convert_pl_monthly()
    verify_pl(pl_output)

    sc_summary, sc_raw = convert_sales_cost()
    verify_sc(sc_summary, pl_output)

    expense_output = convert_expense_detail()
    verify_expense(expense_output, pl_output)

    # KPI 추가 데이터
    convert_kpi_ch()
    convert_kpi_health()
    convert_kpi_nabota()
